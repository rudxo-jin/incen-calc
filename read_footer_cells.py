import openpyxl

try:
    wb = openpyxl.load_workbook('근로소득_간이세액표(조견표).xlsx')
    sheet = wb.active
    
    # Get max row
    max_row = sheet.max_row
    start_row = max(1, max_row - 20)
    
    print(f"Reading from row {start_row} to {max_row}")
    
    for row in sheet.iter_rows(min_row=start_row, max_row=max_row):
        for cell in row:
            if cell.value:
                print(f"[{cell.coordinate}] {cell.value}")
            
except Exception as e:
    print(f"Error: {e}")
