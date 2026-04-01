import openpyxl

try:
    wb = openpyxl.load_workbook('근로소득_간이세액표(조견표).xlsx')
    sheet = wb.active
    
    for r in range(651, 663):
        c1 = sheet.cell(row=r, column=1).value
        c2 = sheet.cell(row=r, column=2).value
        print(f"R{r}: [C1] {c1} [C2] {c2}")
            
except Exception as e:
    print(f"Error: {e}")
